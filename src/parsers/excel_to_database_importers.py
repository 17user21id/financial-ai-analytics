"""
Excel to Database Importers
Read Excel files and import accounts and transactions to database
Handles unique account creation and transaction assignment based on is_summary flag
"""

# Required imports
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
import re

# Required dependencies - fail fast if not available
try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel operations. Please install it with: pip install openpyxl"
    )

from ..stores.database_manager import get_account_store, get_transaction_store
from ..common.enums import DataSource, AccountType, RevenueSubType, DerivedSubType


class AccountManager:
    """
    Manages unique account creation and transaction assignment
    Ensures accounts are created only once and transactions are assigned correctly
    """

    def __init__(self):
        self.account_store = get_account_store()
        self.transaction_store = get_transaction_store()

        # Cache for account lookups to avoid repeated database queries
        self._account_cache: Dict[str, int] = {}  # name -> account_id mapping

    def get_or_create_account(
        self,
        name: str,
        category_path: str,
        sub_category: str,
        account_type: int,
        sub_type: Optional[int],
        is_summary: bool,
        is_derived: bool,
        description: str,
    ) -> Optional[int]:
        """
        Get existing account or create new one if unique

        Args:
            name: Account name
            category_path: Category path
            sub_category: Sub category
            account_type: Account type (1-5)
            sub_type: Sub type (1-7 or None)
            is_summary: Whether account is summary
            is_derived: Whether account is derived
            description: Account description/metadata

        Returns:
            Account ID (None if is_summary is True)
        """
        # Don't create accounts if is_summary is True
        if is_summary:
            return None

        # Check cache first
        cache_key = f"{name}_{category_path}_{account_type}_{sub_type}"
        if cache_key in self._account_cache:
            return self._account_cache[cache_key]

        # Check if account exists in database using composite key
        existing_account = self.account_store.get_account_by_composite_key(
            name, category_path, account_type, sub_type
        )

        if existing_account:
            self._account_cache[cache_key] = existing_account["account_id"]
            return existing_account["account_id"]

        # Create new account
        account_data = {
            "name": name,
            "category_path": category_path,
            "sub_category": sub_category,
            "type": account_type,
            "sub_type": sub_type,
            "is_summary": is_summary,
            "is_derived": is_derived,
            "description": description,
            "is_active": True,
        }

        account_id = self.account_store.create_account(account_data)
        self._account_cache[cache_key] = account_id

        return account_id

    def create_transaction(
        self,
        account_id: int,
        period_start: str,
        period_end: str,
        value: float,
        currency: int = 1,
        derived_sub_type: Optional[int] = None,
        source_id: int = 1,
        notes: str = "",
    ) -> int:
        """
        Create a transaction for the given account

        Args:
            account_id: Account ID
            period_start: Period start date
            period_end: Period end date
            value: Transaction value
            currency: Currency code
            derived_sub_type: Derived sub type
            source_id: Source ID
            notes: Transaction notes

        Returns:
            Transaction ID
        """
        transaction_data = {
            "account_id": account_id,
            "period_start": period_start,
            "period_end": period_end,
            "value": value,
            "currency": currency,
            "derived_sub_type": derived_sub_type,
            "source_id": source_id,
            "notes": notes,
        }

        transaction_id = self.transaction_store.create_transaction(transaction_data)
        return transaction_id

    def clear_cache(self):
        """Clear the account cache"""
        self._account_cache.clear()


class BaseExcelImporter(ABC):
    """
    Base class for importing Excel files to database
    """

    def __init__(self):
        self.account_manager = AccountManager()

        # Account type mapping (string to enum)
        self.account_type_map = {
            "revenue": AccountType.REVENUE,
            "cogs": AccountType.COGS,
            "expense": AccountType.EXPENSE,
            "tax": AccountType.TAX,
            "derived": AccountType.DERIVED,
        }

        # Sub type mapping (string to enum)
        self.sub_type_map = {
            "operating": RevenueSubType.OPERATING,
            "non_operating": RevenueSubType.NON_OPERATING,
            "gross_profit": DerivedSubType.GROSS_PROFIT,
            "operating_profit": DerivedSubType.OPERATING_PROFIT,
            "ebitda": DerivedSubType.EBITDA,
            "non_operating_income": DerivedSubType.NON_OPERATING_INCOME,
            "profit_before_tax": DerivedSubType.PROFIT_BEFORE_TAX,
            "net_profit": DerivedSubType.NET_PROFIT,
            "other": DerivedSubType.OTHER,
        }

    def import_excel_file(self, excel_file_path: str) -> Tuple[int, int]:
        """
        Import Excel file to database

        Args:
            excel_file_path: Path to Excel file

        Returns:
            Tuple of (accounts_created, transactions_created)
        """
        try:
            workbook = load_workbook(excel_file_path)
            accounts_created = 0
            transactions_created = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_accounts, sheet_transactions = self._import_sheet(
                    sheet, sheet_name
                )
                accounts_created += sheet_accounts
                transactions_created += sheet_transactions

            return accounts_created, transactions_created

        except Exception as e:
            raise

    @abstractmethod
    def _import_sheet(self, sheet, sheet_name: str) -> Tuple[int, int]:
        """Import a single sheet - to be implemented by subclasses"""
        pass

    def _map_account_type(self, type_str: str) -> int:
        """Map account type string to enum value"""
        return self.account_type_map.get(type_str.lower(), AccountType.DERIVED).value

    def _map_sub_type(self, sub_type_str: str) -> Optional[int]:
        """Map sub type string to enum value"""
        if not sub_type_str:
            return None
        mapped_type = self.sub_type_map.get(sub_type_str.lower())
        return mapped_type.value if mapped_type else None

    def _parse_headers(self, sheet) -> Dict[str, int]:
        """Parse headers from sheet and return column mapping"""
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))

        header_map = {}
        for i, header in enumerate(headers):
            header_map[header] = i + 1

        return header_map

    def _extract_account_data(
        self, sheet, row_num: int, header_map: Dict[str, int]
    ) -> Optional[Dict[str, Any]]:
        """Extract account data from a row"""
        try:
            name = sheet.cell(row=row_num, column=header_map.get("name", 0)).value
            if not name:
                return None

            return {
                "name": str(name),
                "category_path": str(
                    sheet.cell(
                        row=row_num, column=header_map.get("category_path", 0)
                    ).value
                    or ""
                ),
                "sub_category": str(
                    sheet.cell(
                        row=row_num, column=header_map.get("sub_category", 0)
                    ).value
                    or ""
                ),
                "type_str": str(
                    sheet.cell(row=row_num, column=header_map.get("type", 0)).value
                    or "derived"
                ),
                "sub_type_str": sheet.cell(
                    row=row_num, column=header_map.get("sub_type", 0)
                ).value,
                "is_summary": bool(
                    sheet.cell(
                        row=row_num, column=header_map.get("is_summary", 0)
                    ).value
                ),
                "is_derived": bool(
                    sheet.cell(
                        row=row_num, column=header_map.get("is_derived", 0)
                    ).value
                ),
                "metadata": str(
                    sheet.cell(row=row_num, column=header_map.get("metadata", 0)).value
                    or ""
                ),
            }
        except Exception as e:
            pass

    def _extract_date_from_header(self, header: str) -> Optional[datetime]:
        """Extract date from header string like 'Jan 2020', 'Feb 2021', etc."""
        try:
            # Pattern to match month abbreviations and years
            pattern = r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})$"
            match = re.match(pattern, header.strip())

            if match:
                month_str, year_str = match.groups()
                month_map = {
                    "Jan": 1,
                    "Feb": 2,
                    "Mar": 3,
                    "Apr": 4,
                    "May": 5,
                    "Jun": 6,
                    "Jul": 7,
                    "Aug": 8,
                    "Sep": 9,
                    "Oct": 10,
                    "Nov": 11,
                    "Dec": 12,
                }
                month = month_map[month_str]
                year = int(year_str)
                return datetime(year, month, 1)  # First day of the month
        except Exception:
            pass
        return None

    def _extract_date_range_from_headers(
        self, headers: List[str]
    ) -> Dict[str, Tuple[str, str]]:
        """Extract date ranges from headers and return column mapping with periods"""
        date_columns = {}

        # We need to work with actual column numbers, not filtered header indices
        # So we'll iterate through the actual sheet columns
        return date_columns

    def _extract_date_range_from_sheet(self, sheet) -> Dict[int, Tuple[str, str]]:
        """Extract date ranges from sheet headers and return column mapping with periods"""
        date_columns = {}

        # Iterate through actual columns in the sheet
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col_num)
            if cell.value:
                header = str(cell.value)
                date_obj = self._extract_date_from_header(header)
                if date_obj:
                    # Create period start and end dates
                    period_start = date_obj.strftime(
                        "%Y-%m-%d"
                    )  # First day of the month

                    # Calculate last day of the same month
                    if date_obj.month == 12:
                        # December -> last day is Dec 31
                        period_end = datetime(date_obj.year, 12, 31).strftime(
                            "%Y-%m-%d"
                        )
                    else:
                        # For other months, get the last day by going to first day of next month and subtracting 1 day
                        next_month = datetime(date_obj.year, date_obj.month + 1, 1)
                        last_day = (next_month - timedelta(days=1)).strftime("%Y-%m-%d")
                        period_end = last_day

                    date_columns[col_num] = (period_start, period_end)

        return date_columns

    def _create_account_from_data(self, account_data: Dict[str, Any]) -> Optional[int]:
        """Create account from extracted data"""
        account_type = self._map_account_type(account_data["type_str"])
        sub_type = self._map_sub_type(account_data["sub_type_str"])

        return self.account_manager.get_or_create_account(
            name=account_data["name"],
            category_path=account_data["category_path"],
            sub_category=account_data["sub_category"],
            account_type=account_type,
            sub_type=sub_type,
            is_summary=account_data["is_summary"],
            is_derived=account_data["is_derived"],
            description=account_data["metadata"],
        )


class PLExcelImporter(BaseExcelImporter):
    """
    Importer for data_set_1 Excel files (Rootfi Report format)
    """

    def __init__(self):
        super().__init__()
        self.source_id = DataSource.PL_REPORT.value  # 1

    def _import_sheet(self, sheet, sheet_name: str) -> Tuple[int, int]:
        """Import Rootfi Report sheet"""
        accounts_created = 0
        transactions_created = 0

        # Parse headers
        header_map = self._parse_headers(sheet)

        # Extract date ranges from sheet headers (using actual column numbers)
        date_columns = self._extract_date_range_from_sheet(sheet)

        # Find value columns (monthly data)
        value_cols = []
        for header, col_num in header_map.items():
            if header not in [
                "name",
                "category_path",
                "sub_category",
                "type",
                "sub_type",
                "is_summary",
                "is_derived",
                "is_derived_correct",
                "metadata",
                "Computed Total",
                "Total",
                "Difference",
            ]:
                value_cols.append(col_num)

        # Validate that we have date information for value columns
        if not date_columns and value_cols:
            raise ValueError(
                f"No date columns found in headers. Headers: {headers}. Expected format: 'Jan 2020', 'Feb 2020', etc."
            )

        # Process each row
        for row_num in range(2, sheet.max_row + 1):
            account_data = self._extract_account_data(sheet, row_num, header_map)
            if not account_data:
                continue

                # Create/get account
            account_id = self._create_account_from_data(account_data)
            if account_id:
                accounts_created += 1

            # Create transactions only for non-summary accounts
            if not account_data["is_summary"]:
                for col in value_cols:
                    value = sheet.cell(row=row_num, column=col).value
                    if value is not None and isinstance(value, (int, float)):
                        # Must have date information for the column
                        if col not in date_columns:
                            raise ValueError(
                                f"No date information found for column {col}. Available date columns: {list(date_columns.keys())}"
                            )

                        period_start, period_end = date_columns[col]

                        self.account_manager.create_transaction(
                            account_id=account_id,
                            period_start=period_start,
                            period_end=period_end,
                            value=float(value),
                            currency=1,
                            source_id=self.source_id,
                            notes=f"Rootfi Report - {sheet_name}",
                        )
                        transactions_created += 1

        return accounts_created, transactions_created

    def _extract_period_from_sheet_name(self, sheet_name: str) -> Tuple[str, str]:
        """Extract period from sheet name"""
        raise ValueError(
            f"Could not extract period from sheet name: '{sheet_name}'. Expected format: 'YYYY-MM-DD to YYYY-MM-DD'"
        )


class RootfiExcelImporter(BaseExcelImporter):
    """
    Importer for data_set_2 Excel files (P&L Report format)
    """

    def __init__(self):
        super().__init__()
        self.source_id = DataSource.ROOTFI_REPORT.value  # 2

    def _import_sheet(self, sheet, sheet_name: str) -> Tuple[int, int]:
        """Import P&L Report sheet"""
        accounts_created = 0
        transactions_created = 0

        # Extract period from sheet name
        period_start, period_end = self._extract_period_from_sheet_name(sheet_name)

        # Parse headers
        header_map = self._parse_headers(sheet)

        # Process each row
        for row_num in range(2, sheet.max_row + 1):
            account_data = self._extract_account_data(sheet, row_num, header_map)
            if not account_data:
                continue

            # Create/get account
            account_id = self._create_account_from_data(account_data)
            if account_id:
                accounts_created += 1

            # Create transaction only for non-summary accounts and non-null values
            if not account_data["is_summary"]:
                value = sheet.cell(row=row_num, column=header_map.get("value", 0)).value
                if value is not None and isinstance(value, (int, float)):
                    self.account_manager.create_transaction(
                        account_id=account_id,
                        period_start=period_start,
                        period_end=period_end,
                        value=float(value),
                        currency=1,
                        source_id=self.source_id,
                        notes=f"P&L Report - {sheet_name}",
                    )
                    transactions_created += 1

        return accounts_created, transactions_created

    def _extract_period_from_sheet_name(self, sheet_name: str) -> Tuple[str, str]:
        """Extract period from sheet name (format: '2022-08-01 to 2022-08-31')"""
        try:
            if " to " in sheet_name:
                start_str, end_str = sheet_name.split(" to ")
                start_str = start_str.strip()
                end_str = end_str.strip()

                # Validate date format
                from datetime import datetime

                datetime.strptime(start_str, "%Y-%m-%d")
                datetime.strptime(end_str, "%Y-%m-%d")

                return start_str, end_str
        except Exception as e:
            pass

        raise ValueError(
            f"Could not extract period from sheet name: '{sheet_name}'. Expected format: 'YYYY-MM-DD to YYYY-MM-DD'. Error: {e}"
        )


class ExcelImporterFactory:
    """
    Factory for creating appropriate Excel importers based on file type
    """

    @staticmethod
    def create_importer(excel_file_path: str) -> BaseExcelImporter:
        """
        Create appropriate importer based on Excel file content

        Args:
            excel_file_path: Path to Excel file

        Returns:
            Appropriate importer instance
        """
        try:
            # First check filename pattern for better detection
            filename = Path(excel_file_path).name.lower()
            if "dataset1" in filename:
                return PLExcelImporter()
            elif "dataset2" in filename:
                return RootfiExcelImporter()

            # Fallback to content-based detection
            workbook = load_workbook(excel_file_path)

            # Check sheet names and content to determine type
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Check if it's a period-based format (data_set_2)
                if " to " in sheet_name and "account_id" in str(sheet[1]):
                    return RootfiExcelImporter()

                # Check if it's Rootfi format (data_set_1)
                if "P&L Report" in sheet_name or "name" in str(sheet[1]):
                    return PLExcelImporter()

            # If no pattern matches, default to PLExcelImporter
            return PLExcelImporter()

        except Exception as e:
            raise Exception(f"Failed to determine Excel importer type: {e}")


# Convenience functions
def import_excel_to_database(excel_file_path: str) -> Tuple[int, int]:
    """
    Import Excel file to database using appropriate importer

    Args:
        excel_file_path: Path to Excel file

    Returns:
        Tuple of (accounts_created, transactions_created)
    """
    importer = ExcelImporterFactory.create_importer(excel_file_path)
    return importer.import_excel_file(excel_file_path)


def convert_and_import_json(
    json_file_path: str, output_excel_path: str = None
) -> Tuple[int, int]:
    """
    Convert JSON to Excel and then import to database

    Args:
        json_file_path: Path to JSON file
        output_excel_path: Optional output Excel path

    Returns:
        Tuple of (accounts_created, transactions_created)
    """
    from .json_to_excel_converters import convert_rootfi_to_excel, convert_pl_to_excel

    # Determine converter based on file name
    if "data_set_1" in json_file_path or "dataset1" in json_file_path:
        if not output_excel_path:
            output_excel_path = json_file_path.replace(".json", "_output.xlsx")
        convert_rootfi_to_excel(json_file_path, output_excel_path)
    else:
        if not output_excel_path:
            output_excel_path = json_file_path.replace(".json", "_output.xlsx")
        convert_pl_to_excel(json_file_path, output_excel_path)

    # Import to database
    return import_excel_to_database(output_excel_path)
